import csv
import os
import re
import time
import random
import subprocess
import pandas as pd
from faker import Faker
from datetime import datetime, timedelta
from playwright.sync_api import Playwright, sync_playwright
from playwright.sync_api import expect
from undetected_playwright import stealth_sync
from openpyxl.styles import PatternFill
import openpyxl

catchall = "gmail.com"


def load_proxies(csv_file):
    """Load proxies from a CSV file and return them as a list."""
    if not os.path.exists(csv_file):
        print(f"Error: The file '{csv_file}' does not exist.")
        return []

    proxies = []
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            if len(row) == 4:  # Ensure row has exactly 4 elements
                proxies.append({
                    "server": f"http://{row[0]}:{row[1]}",
                    "username": row[2],
                    "password": row[3]
                })
            else:
                print(f"Skipping invalid row: {row}")  # Log invalid rows
    return proxies

def assign_proxies_to_accounts(accounts, proxies):
    """Assign proxies to accounts when they are created or first logged into."""
    assigned_proxies = {}

    # Assign proxies to newly created accounts
    for i, account in enumerate(accounts):
        if "proxy" not in account or account["proxy"] is None:  # Assign only if not already set
            proxy_index = i % len(proxies)  # Rotate proxies if more accounts than proxies
            assigned_proxies[account["email"]] = proxies[proxy_index]

    print(f"[DEBUG] Assigned Proxies: {assigned_proxies}")  # Debugging to check proxy assignment
    return assigned_proxies

def apply_conditional_formatting(file_name):
    """Apply conditional formatting to the Status column."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    # Find the "Status" column
    status_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Status":
            status_col = idx
            break

    if status_col:
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

        for row in ws.iter_rows(min_row=2, min_col=status_col, max_col=status_col):
            for cell in row:
                if cell.value == "Active":
                    cell.fill = green_fill
                elif cell.value == "Not Active":
                    cell.fill = red_fill

    wb.save(file_name)


def log_stats_to_excel(account_number, email, password, success, alr_processed, points=None, status="Active"):
    """Log or update stats in the account stats Excel file."""
    file_name = "account_stats.xlsx"
    columns = ["Account Number", "Email", "Password", "Success", "Already Processed", "Points", "Status", "Frequency Count"]

    # Check if file exists and load data
    if os.path.exists(file_name):
        df = pd.read_excel(file_name)
        if not all(col in df.columns for col in columns):
            print(f"Warning: The file '{file_name}' has incorrect or missing columns. Recreating it.")
            df = pd.DataFrame(columns=columns)
    else:
        df = pd.DataFrame(columns=columns)

    # Update existing stats or add new ones
    if email in df["Email"].values:
        current_freq_count = df.loc[df["Email"] == email, "Frequency Count"].fillna(1).iloc[0]
        current_status = df.loc[df["Email"] == email, "Status"].iloc[0]

        if points is not None:
            current_points = df.loc[df["Email"] == email, "Points"].fillna(0).iloc[0]
            point_difference = points - current_points

            # **🔹 Check for inactivity**
            if point_difference > 500:
                status = "Not Active"
            elif point_difference == 0:
                alr_processed += 1  # Increment already processed
                status = "Not Active"
            elif point_difference > 1:
                success += 1  # Increment success
            else:
                status = "Active"   

            # **🔹 Track frequency count properly**
            if current_status == "Active":
                current_freq_count += 1  # Increment only if active

            # **🔹 Make account inactive after 3 cycles**
            if current_freq_count >= 3 and status == "Active":
                print(f"[INFO] Account {email} reached max cycles. Setting to Not Active.")
                current_freq_count = 0
                status = "Not Active"  
            elif current_status == "Not Active":
                print(f"[INFO] Account {email} is inactive, keeping it inactive.")
                status = "Not Active"  # Keep it inactive

            # **🔹 Update Points, Status, and Frequency Count**
            df.loc[df["Email"] == email, "Points"] = points
            df.loc[df["Email"] == email, "Status"] = status
            df.loc[df["Email"] == email, "Frequency Count"] = current_freq_count

        # **🔹 Update success & already processed counts**
        df.loc[df["Email"] == email, "Success"] += success
        df.loc[df["Email"] == email, "Already Processed"] += alr_processed

    else:
        # Add a new row for accounts not already in the DataFrame
        new_row = {
            "Account Number": account_number,
            "Email": email,
            "Password": password,
            "Success": success,
            "Already Processed": alr_processed,
            "Points": points,
            "Status": status,
            "Frequency Count": 1  # Start at 1
        }
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

    # **🔹 Save updates to Excel**
    df.to_excel(file_name, index=False)

    # **🔹 Apply conditional formatting**
    apply_conditional_formatting(file_name)



def escape_latex(text):
    """Escape LaTeX special characters."""
    return text.replace("#", "\\#")

def generate_random_receipt():
    """Generate random receipt details."""
    tc_number = escape_latex(f"TC# {random.randint(1000, 9999)} {random.randint(1000, 9999)} {random.randint(1000, 9999)} {random.randint(1000, 9999)}")
    st_number = escape_latex(f"ST# {random.randint(1000, 9999)} OP# {random.randint(1000, 9999)} TE# {random.randint(1, 20)} TR# {random.randint(1000, 9999)}")
    random_date = (datetime.now() - timedelta(days=random.randint(0, 15))).strftime("%m/%d/%y %H:%M:%S")
    amex_number = escape_latex(f"{random.randint(1000, 9999)}")

    # Item names and numbers
    items_dict = {
        "OILSPRAY": "002639599991 F",
        "PLSBY ELF": "001800011925 F",
        "BEEF RIBEYE": "026039400000 F",
        "CHILL 12PK": "004900055539 F",
        "CAKE": "007432309524 F",
        "B J ICECRM": "007684040007 F",
        "GV HF HF": "060538818715 F",
        "GV CK MJ 8Z": "007874203972 F"
    }

    # Fixed MMZ Lemonade
    # Make a matrix for sporte packages (6-packs, 2-liters)
    fixed_item = ("MMZ LEMONADE", "002500012052 F", random.uniform(2.50,4.50))

    # Randomly select 4 items excluding MMZ Lemonade
    selected_items = random.sample(list(items_dict.items()), 4)

    # Generate random prices for selected items
    items_with_prices = [(name, number, round(random.uniform(10.0, 30.0), 2)) for name, number in selected_items]
    
    # changing variable for spacing when subtotal is over 99.99 
    # Combine fixed MMZ Lemonade with the random items 
    # Take sporte out of items and add underneath to fix 3 decimal spacing
    items = items_with_prices[:2] + [fixed_item] + items_with_prices[2:]

    subtotal = round(sum(price for _, _, price in items), 2)
    tax1 = round(0.07 * subtotal, 2)
    total = round(subtotal + tax1, 2)

    return tc_number, st_number, random_date, amex_number, items, subtotal, tax1, total

def create_receipt_latex(tc_number, st_number, random_date, amex_number, items, subtotal, tax1, total, logo_path, barcode_path):
    """Create LaTeX receipt."""
    items_tabbed = r"""
    \begin{tabbing}
        \hspace{2.5cm} \= \hspace{3.7cm} \= \kill
"""
    for name, number, price in items:
        items_tabbed += f"        \\textbf{{{name}}} \\> \\textbf{{{number}}} \\> \\textbf{{{price:.2f}}}\\\\\n"
    items_tabbed += r"    \end{tabbing}"

    receipt_template = r"""
\documentclass{article}
\usepackage{geometry}
\geometry{paperwidth=100mm, paperheight=150mm, left=5mm, top=5mm, right=5mm, bottom=5mm}
\usepackage{courier}
\renewcommand{\familydefault}{\ttdefault}
\usepackage{array}
\usepackage{graphicx}
\usepackage{multicol}
\usepackage{xcolor}
\pagecolor{white}
\usepackage{adjustbox}

\begin{document}
\pagestyle{empty}

\newcommand{\receiptfontsize}{\fontsize{10}{9}\selectfont}
\receiptfontsize

\begin{center}
    \includegraphics[width=\linewidth]{""" + logo_path + r"""} % Walmart logo
    \textbf{WAL*MART}\\
    \textbf{33062990 Mgr. MIRANDA}\\
    \textbf{905 SINGLETARY DR}\\
    \textbf{STREETSBORO, OH}\\
    \textbf{""" + st_number + r"""}\\
    
    \vspace{1mm}
""" + items_tabbed + r"""
    \vspace{-9mm}
    \hspace{2.5cm}\begin{tabbing}
        \hspace{2cm} \= \hspace{2.4cm} \= \kill
        \textbf{\hspace{4cm}SUBTOTAL} \> \textbf{\hspace{4.3cm} """ + str(subtotal) + r"""} \\
        \textbf{\hspace{2.5cm}TAX} \hspace{-0.25mm} \textbf{1} \> \textbf{\hspace{2.5cm}7\%} \> \textbf{\hspace{1.90cm} """ + str(tax1) + r"""} \\
        \textbf{\hspace{2.5cm}TAX 12} \> \textbf{\hspace{2.5cm}0\%} \> \textbf{\hspace{2.1cm}0.00} \\
        \textbf{\hspace{4.75cm}TOTAL} \> \textbf{\hspace{4.3cm} """ + str(total) + r"""} \\
        \textbf{\hspace{2.45cm}AMEX CREDIT TEND} \> \textbf{\hspace{4.3cm} """ + str(total) + r"""} \\
        \textbf{\hspace{2.7cm}AMEX} \textbf{**** **** **** """ + amex_number + r"""} \\
        \textbf{\hspace{3.7cm}CHANGE DUE} \> \textbf{\hspace{4.7cm}0.00} \\
    \end{tabbing}
    
    \vspace{-2mm}

    \textbf{\huge{\# ITEMS SOLD 5}}\\
    \vspace{0.5cm}
    \textbf{""" + tc_number + r"""}\\
    \includegraphics[width=\linewidth]{""" + barcode_path + r"""} % Barcode
    \vspace{0.5cm}
    \textbf{""" + random_date + r"""}
\end{center}  
\end{document}
"""
    with open("receipt.tex", "w") as f:
        f.write(receipt_template)

def compile_latex_to_png():
    """Compile LaTeX to PDF and convert to PNG."""
    try:
        subprocess.run(["pdflatex", "receipt.tex"], check=True)
        subprocess.run(["convert", "-density", "200", "receipt.pdf", "-quality", "100", "receipt.png"], check=True)
    except subprocess.CalledProcessError as e:
        print(f"Error during compilation: {e}")
        return False
    return True

def create_account(playwright, accounts, account_number, proxies, assigned_proxies):
    """Create an account on Cavs Rewards and assign a proxy to it."""
    
    # ✅ Generate the email FIRST before assigning a proxy
    email = f"{Faker().last_name()}{random.randint(1000, 9999)}@{catchall}"
    password = f"Pass{random.randint(1000, 9999)}{Faker().word()}!"

    # ✅ Assign a proxy AFTER email is created
    proxy_index = len(accounts) % len(proxies)  
    assigned_proxies[email] = proxies[proxy_index]  # Now 'email' exists

    proxy_config = assigned_proxies[email]  # Use assigned proxy

    # ✅ Debugging print statement to check proxy assignment
    print(f"[INFO] Creating Account {email} using Proxy: {proxy_config}")
    
     # Use assigned proxy
    browser = playwright.chromium.launch(
        headless=False,
        proxy=proxy_config
    )
    context = browser.new_context()
    page = context.new_page()

    # Apply stealth to avoid bot detection
    stealth_sync(page)

    page.goto("https://www.cavsrewards.com/")
    time.sleep(.7)
    page.get_by_placeholder("Referral Code (optional)").click()
    time.sleep(0.7)
    page.get_by_placeholder("Referral Code (optional)").fill("BJjPn")
    time.sleep(0.7)
    page.get_by_role("button", name="Continue to Cavs Rewards").click()
    time.sleep(0.7)
    page.get_by_role("link", name="Create account now").click()

    email = f"{Faker().last_name()}{random.randint(1000, 9999)}@{catchall}"
    password = f"Pass{random.randint(1000, 9999)}{Faker().word()}!"

    time.sleep(0.7)
    page.get_by_label("Email address").click()
    time.sleep(0.7)
    page.get_by_label("Email address").fill(email)
    time.sleep(0.7)
    page.get_by_label("Password").click()
    time.sleep(0.7)
    page.get_by_label("Password").fill(password)
    time.sleep(0.7)
    page.get_by_role("button", name="Continue", exact=True).click()
    time.sleep(0.7)
    page.get_by_role("button", name="GET STARTED").click()
    time.sleep(0.7)
    page.get_by_role("button", name="Continue").click()
    time.sleep(0.7)
    page.get_by_role("button", name="Continue").click()
    time.sleep(0.7)
    page.get_by_role("button", name="Continue").click()
    time.sleep(0.7)
    page.get_by_role("button", name="Continue").click()
    time.sleep(0.7)
    page.locator("svg").click()
    time.sleep(0.7)
    page.get_by_role("button", name="Done").click()

    # Add success and alr_processed counters to the account details
    account_details = {
        "account_number": account_number,
        "email": email,
        "password": password,
        "success": 0,
        "alr_processed": 0
    }
    accounts.append(account_details)

    log_stats_to_excel(account_number, email, password, 0, 0)

    context.close()
    browser.close()
    print(f"Account created: {email}")

def login_and_upload_receipt(playwright, account, receipt_path, assigned_proxies, proxies):
    """Log into an account and upload receipt."""
    email = account["email"]
    
    # Assign a proxy if this account doesn't have one yet
    if email not in assigned_proxies:
        proxy_index = len(assigned_proxies) % len(proxies)  # Assign sequentially
        assigned_proxies[email] = proxies[proxy_index]

    proxy_config = assigned_proxies[email]

    browser = playwright.chromium.launch(headless=False, proxy=proxy_config)
    context = browser.new_context()
    page = context.new_page()

    # Ensure Stealth is applied to bypass detection
    stealth_sync(page) 

    # Load account stats
    file_name = "account_stats.xlsx"
    df = pd.read_excel(file_name)
    account_status = df.loc[df["Email"] == account["email"], "Status"].iloc[0]

    # If the account is not active, set it to active and skip this iteration
    if account_status == "Not Active":
        print(f"Account {account['email']} is not active. Resetting status to Active and skipping.")
        df.loc[df["Email"] == account["email"], "Status"] = "Active"
        df.to_excel(file_name, index=False)
        context.close()
        browser.close()
        return
    
    page.goto("https://www.cavsrewards.com/auth")
    page.get_by_placeholder("Referral Code (optional)").click()
    time.sleep(0.7)
    page.get_by_placeholder("Referral Code (optional)").fill("BJjPn")
    time.sleep(0.7)
    page.get_by_role("button", name="Continue to Cavs Rewards").click()
    time.sleep(0.7)
    page.get_by_label("Email address").fill(account["email"])
    time.sleep(0.7)
    page.get_by_label("Password").fill(account["password"])
    time.sleep(0.7)
    page.get_by_role("button", name="Continue", exact=True).click()
    time.sleep(0.7)

    try:
        expect(page.get_by_role("button", name="GET STARTED")).to_be_visible(timeout=2000)
        print("[INFO] 'GET STARTED' button found! Clicking now...")
        page.get_by_role("button", name="GET STARTED").click()
        time.sleep(0.7)
        page.get_by_role("button", name="Continue").click()
        time.sleep(0.7)
        page.get_by_role("button", name="Continue").click()
        time.sleep(0.7)
        page.get_by_role("button", name="Continue").click()
        time.sleep(0.7)
        page.get_by_role("button", name="Continue").click()
        time.sleep(0.7)
        page.locator("svg").click()
        time.sleep(0.7)
        page.get_by_role("button", name="Done").click()
    except:
        print("[WARNING] 'GET STARTED' not found. Continuing without it.")
        
    page.get_by_role("img", name="Close").click()
    time.sleep(0.7)
    page.get_by_role("link", name="Card Top Coca-Cola Products").click()
    time.sleep(0.7)
    page.locator('input[type="file"]').set_input_files(receipt_path)
    time.sleep(0.7)
    page.get_by_role("button", name="Check").click()
    time.sleep(15)
    page.get_by_role("img", name="Close").click()
    time.sleep(0.7)
    page.get_by_role("link", name="Rewards").click()
    time.sleep(8)
    # Check if the text "Lifetime:" exists in the body
    body_text = page.locator("body").text_content()
    assert "Lifetime:" in page.locator("body").text_content(), "Text 'Lifetime:' not found on the page"
    # Retrieve the text content of the <body>
    # Extract points
    points = None
    if "Lifetime:" in body_text:
        matches = re.findall(r"Lifetime:\s*([\d,]+)", body_text)
        if matches:
            points = int(matches[0].replace(",", ""))
            print(f"Account points for {account['email']}: {points}")

    # Update stats
    log_stats_to_excel(
        account_number=account["account_number"],
        email=account["email"],
        password=account["password"],
        success=account["success"],
        alr_processed=account["alr_processed"],
        points=points
    )
    
    context.close()
    browser.close()

def main():
    """Main function to create or log into accounts with assigned proxies."""
    proxies = load_proxies("proxies.csv")
    file_name = "account_stats.xlsx"
    accounts = []

    # Load existing accounts
    if os.path.exists(file_name):
        df = pd.read_excel(file_name)
        if not df.empty:
            for _, row in df.iterrows():
                accounts.append({
                    "account_number": row["Account Number"],
                    "email": row["Email"],
                    "password": row["Password"],
                    "success": row["Success"],
                    "alr_processed": row["Already Processed"]
                })
    
    # Assign proxies (either at creation or when logging in)
    assigned_proxies = assign_proxies_to_accounts(accounts, proxies)

    with sync_playwright() as playwright:
        # If no accounts exist, create them
        if not accounts:
            print("[INFO] No existing accounts found. Creating new accounts...")
            for i in range(1, 6):  # Create up to 5 accounts
                create_account(playwright, accounts, i, proxies, assigned_proxies)
                time.sleep(random.uniform(2, 5))

        else:
            print("[INFO] Using existing accounts for login and receipt upload.")

        while True:
            for account in accounts:
                email = account["email"]
                print(f"[INFO] Processing account {email}...")

                if email not in assigned_proxies:
                    print(f"[ERROR] No proxy assigned for {email}. Assigning now...")
                    assigned_proxies[email] = proxies[len(assigned_proxies) % len(proxies)]

                # Generate and upload receipt
                tc_number, st_number, random_date, amex_number, items, subtotal, tax1, total = generate_random_receipt()
                create_receipt_latex(tc_number, st_number, random_date, amex_number, items, subtotal, tax1, total, "Header.png", "barcode.png")

                if compile_latex_to_png():
                    receipt_path = "receipt.png"
                    login_and_upload_receipt(playwright, account, receipt_path, assigned_proxies, proxies)
                else:
                    print(f"[ERROR] Failed to generate receipt for {email}. Skipping.")

                time.sleep(random.uniform(5, 15))

if __name__ == "__main__":
    main()
